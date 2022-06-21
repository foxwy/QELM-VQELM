# -*- coding: utf-8 -*-

import openpyxl
import numpy as np

def Cal_mean_std(R):
    R.sort()
    R_data = R#[5:95]
    R_mean = np.mean(R_data)
    R_std = np.std(R_data, ddof=1)

    return R_mean, R_std


def Plot_data(flag=1):
    if flag == 1:
        wb = openpyxl.load_workbook('figure_final/sqrt_noise/results.xlsx')
    else:
        wb = openpyxl.load_workbook('results.xlsx')
    ws = wb.get_sheet_by_name('Sheet')

    r1_mean = []
    r1_std = []
    r2_mean = []
    r2_std = []
    r3_mean = []
    r3_std = []
    r4_mean = []
    r4_std = []
    tr1_mean = []
    tr2_mean = []
    tr3_mean = []
    tr4_mean = []
    print('\n')
    for k in range(1, 7):
        R1 = []
        R2 = []
        R3 = []
        R4 = []
        Tr1 = []
        Te1 = []
        Tr2 = []
        Te2 = []
        Tr3 = []
        Te3 = []
        Tr4 = []
        Te4 = []
        j = 100 * (k - 1) + 1
        for r1, r2, r3, r4, tr1, tr2, tr3, tr4 in zip(ws['D'][j:j + 100], ws['E'][j:j + 100], ws['F'][j:j + 100], ws['G'][j:j + 100], 
                                  ws['H'][j:j + 100], ws['J'][j:j + 100], 
                                  ws['L'][j:j + 100], ws['N'][j:j + 100]):
            R1.append(r1.value)
            R2.append(r2.value)
            R3.append(r3.value)
            R4.append(r4.value)
            Tr1.append(tr1.value)
            Tr2.append(tr2.value)
            Tr3.append(tr3.value)
            Tr4.append(tr4.value)

        R1_mean, R1_std = Cal_mean_std(R1)
        R2_mean, R2_std = Cal_mean_std(R2)
        R3_mean, R3_std = Cal_mean_std(R3)
        R4_mean, R4_std = Cal_mean_std(R4)
        Tr1_mean, _ = Cal_mean_std(Tr1)
        Tr2_mean, _ = Cal_mean_std(Tr2)
        Tr3_mean, _ = Cal_mean_std(Tr3)
        Tr4_mean, _ = Cal_mean_std(Tr4)
        print('%s & %s & %s & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e \\\\' % (ws['A'][j].value, ws['B'][j].value, ws['C'][j].value, 
                                                                        R1_mean, R2_mean, R3_mean, R4_mean,
                                                                        R1_std, R2_std, R3_std, R4_std,
                                                                        Tr1_mean, Tr2_mean, Tr3_mean, Tr4_mean))
        r1_mean.append(R1_mean)
        r1_std.append(R1_std)
        r2_mean.append(R2_mean)
        r2_std.append(R2_std)
        r3_mean.append(R3_mean)
        r3_std.append(R3_std)
        r4_mean.append(R4_mean)
        r4_std.append(R4_std)
        tr1_mean.append(Tr1_mean)
        tr2_mean.append(Tr2_mean)
        tr3_mean.append(Tr3_mean)
        tr4_mean.append(Tr4_mean)
    print('%.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e \\\\' % (np.mean(r1_mean), np.mean(r2_mean), np.mean(r3_mean), np.mean(r4_mean),
                                                                         np.mean(r1_std), np.mean(r2_std), np.mean(r3_std), np.mean(r4_std),
                                                                         np.mean(tr1_mean), np.mean(tr2_mean), np.mean(tr3_mean), np.mean(tr4_mean)))

if __name__ == '__main__':
    Plot_data(2)