# -*- coding: utf-8 -*-

import openpyxl
import numpy as np

def Cal_mean_std(R):
    R.sort()
    R_data = R#[5:95]
    R_mean = np.mean(R_data)
    R_std = np.std(R_data, ddof=1)

    return R_mean, R_std


def Plot_data(flag=1):
    if flag == 1:
        wb = openpyxl.load_workbook('figure_final/concrete/results.xlsx')
    else:
        wb = openpyxl.load_workbook('results.xlsx')
    ws = wb.get_sheet_by_name('Sheet')

    R1 = []
    R2 = []
    R3 = []
    R4 = []
    Tr1 = []
    Tr2 = []
    Tr3 = []
    Tr4 = []
    j = 1
    for r1, r2, r3, r4, tr1, tr2, tr3, tr4 in zip(ws['A'][j:j + 10], ws['B'][j:j + 10], ws['C'][j:j + 10], ws['D'][j:j + 10], 
                              ws['E'][j:j + 10], ws['G'][j:j + 10], 
                              ws['I'][j:j + 10], ws['K'][j:j + 10]):
        R1.append(r1.value)
        R2.append(r2.value)
        R3.append(r3.value)
        R4.append(r4.value)
        Tr1.append(tr1.value)
        Tr2.append(tr2.value)
        Tr3.append(tr3.value)
        Tr4.append(tr4.value)

    R1_mean, R1_std = Cal_mean_std(R1)
    R2_mean, R2_std = Cal_mean_std(R2)
    R3_mean, R3_std = Cal_mean_std(R3)
    R4_mean, R4_std = Cal_mean_std(R4)
    Tr1_mean, _ = Cal_mean_std(Tr1)
    Tr2_mean, _ = Cal_mean_std(Tr2)
    Tr3_mean, _ = Cal_mean_std(Tr3)
    Tr4_mean, _ = Cal_mean_std(Tr4)
    print('\n%.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e & %.2e \\\\' % (R1_mean, R2_mean, R3_mean, R4_mean,
                                                                    R1_std, R2_std, R3_std, R4_std,
                                                                    Tr1_mean, Tr2_mean, Tr3_mean, Tr4_mean))

if __name__ == "__main__":
    Plot_data(2)